import query
import openpyxl

# comb = [] #Stands for COMBINACIONES
# l_comb = [] #Stands for LOCALIZACION DE COMBINACIONES
reng_enc = 2 #Stands for renglones de encabezados, renglones que no son necesario checar, 
			 # un renglon para pasar del marcador al encabezado (1) y otro para pasar del encabezado al primer item (2)
# respuesta = []

# def add_comb():
	# print ('¿Deseas agregar otra combinacion?, Escribela a continuación, de lo contrario presiona la tecla TAB')
	# respuesta = input()
	# if respuesta != '\t':
		# comb.append(respuesta)
		# add_comb()
	# else:
		# print ('> Localizando combinaciones en archivo de planeación')

archivo = input("¿Cual es el archivo a actualizar?\npor favor, toma en cuenta que la hoja default es [Sheet1]\n>")
sheet = "Sheet1"

# print ("Escribe la(s) combinacion(es) que se cotejaran\n>")
# comb.append(input())
# add_comb()

# print (f"Buscando en archivo \"{archivo}\" con hoja de calculo {sheet}")
# print (f"buscando variables {comb}")

# for i in range(len(comb)):
	# a = query.finding(archivo, sheet, comb[i] )
	# l_comb.append(a)
	
archivo_busqueda = openpyxl.load_workbook(archivo)
ws = archivo_busqueda[sheet]
a = ws.cell(127,1)
c = a.row + reng_enc

# for i in range (1, 50, 1):
	# b = ws.cell(a.row + i,4)
	# print (b.value)
d = a.value
# contador = 0
i = 0
while d != None:
	b = ws.cell(row = c + i, column = 3)
	print (b.value)
	d= b.value
	i += 1
	
print("He acabado señor ,desea algo más?")

#Buscado "2022 V363" la salida " 2022 V363, numeros cuando sean numeros, "