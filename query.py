import openpyxl
import arrow


def finding(file, xlsheet, word):
    ''' Te dice en que celda se encuentra el valor buscado '''
    # Abre Archivo
    archivo = openpyxl.load_workbook(file)
    hoja = archivo[xlsheet]
    goal = []
    # Barre desde la celda A1 hasta la ultima escrita compara
    # los valores con la palabra y retorna el valor de la celda
    for i in range (1, hoja.max_row, 1):
        for j in range (1, hoja.max_column, 1):
            info = hoja.cell(row = i, column = j, )
            if info.value == word:
                goal = hoja.cell(row = i, column = j, )
    return goal
	
def testo():
	a = "me la pelas"
	return (a)

# '''Al final lo se se puede hacer en grupo, se puede hacer por uno solo'''
# ''' Obtiene una celda resultado de la intersección de dos valores dados
#      Para  el master plan ayuda a obtener una fecha del milestone de programa deseado '''
# def intersection_single(file, xlsheet, row_word, column_word):
#     ''' Abre Archivo '''
#     archivo = openpyxl.load_workbook(file)
#     hoja = archivo[xlsheet]
#     ''' Encuentra las celdas '''
#     goal_row = finding(file, xlsheet, row_word)
#     goal_column = finding(file, xlsheet, column_word)
#     ''' Realice el cruce'''
#     goal1 = hoja.cell(row = goal_row.row, column = goal_column.column)
#     return goal1


def intersection_group(file, xlsheet, row_word, column_group):
    ''' Obtiene un grupo de celdas resultado de la busqueda de la intersección de dos valores dados
    Para  el master plan ayuda a obtener varias fecha del milestone de programa deseado '''
    # Abre Archivo
    archivo = openpyxl.load_workbook(file)
    hoja = archivo[xlsheet]
    goal1 = [0]*len(column_group)

    # Encuentra la celda de la fila
    for i in range (1, hoja.max_row, 1):
        for j in range (1, hoja.max_column, 1):
            info = hoja.cell(row = i, column = j, )
            if info.value == row_word:
                goal_row = hoja.cell(row = i, column = j, )

    # Conforme encuentra las celdas de las columnas, realiza el cruce
    for k in range (len(column_group)):
        for i in range (1, hoja.max_row, 1):
            for j in range (1, hoja.max_column, 1):
                info = hoja.cell(row = i, column = j, )
                if info.value == column_group[k]:
                    goal_column = hoja.cell(row = i, column = j, )
        goal1[k] = hoja.cell(row = goal_row.row, column = goal_column.column)#Si te sale error si es el proyecto correto?#
    return goal1



def find_rep_values(file, o_xlsheet, d_xlsheet_name, reference, values):
    '''find & replace values, reference should be in format x=[, ,]'''
    # Abre Archivo
    archivo = openpyxl.load_workbook(file)
    template = archivo[o_xlsheet]
    target = archivo.copy_worksheet(template)
    target.title = d_xlsheet_name
    hoja = archivo[d_xlsheet_name]

    reference = [reference]
    for k in range (len(reference)):
        for i in range (1, hoja.max_row + 1, 1):
            for j in range (1, hoja.max_column + 1, 1):
                celda = hoja.cell(row = i, column = j, )
                # print(celda.value)
                if celda.value != None:
                    info = celda.value.split()# Error: ¿archivo tiene un espacio? #
                    # print (info)
                    for h in range (len(info)):
                        if info[h] == reference[k]:
                            info [h] = values
                            # print (celda.value, info)
                            for m in range ((len(info)) - 1):
                                info [m + 1] = info[m] + ' ' + info [m + 1]
                                data = target.cell(row = i, column = j, value = info[((len(info))-1)] )

    archivo.save(file)


def find_rep_dates(file, o_xlsheet, d_xlsheet_name, reference, values):
    '''find & replace dates, values should be in format x=[, ,]'''
    archivo = openpyxl.load_workbook(file)
    template = archivo[o_xlsheet]
    target = archivo.copy_worksheet(template)
    target.title = d_xlsheet_name
    hoja = archivo[d_xlsheet_name]

    # Conforme encuentra las celdas de las columnas, realiza el cambio
    for k in range (len(reference)):
        for i in range (1, hoja.max_row + 1, 1):
            for j in range (1, hoja.max_column + 1, 1):
                celda = hoja.cell(row = i, column = j, )
                # print(celda)
                if celda.value != None:
                    info = celda.value.split()# Error: ¿archivo tiene un espacio? #
                    if (len(info)) == 1 and info[0] == reference[k]:
                        fecha = values[k]
                        hoja.cell(row = i, column = j, value = fecha.format('DD/MM/YYYY'))

                    if (len(info)) == 3 and info[0] == reference[k]:
                        info[0] = values[k]
                        if info[2] == 'weeks' or info[2] == 'w':
                            fecha = info[0].shift(weeks=float(info[1]))
                            hoja.cell(row = i, column = j, value = fecha.format('DD/MM/YYYY'))

                    if (len(info)) == 3 and info[0] == reference[k]:
                        info[0] = values[k]
                        if info[2] == 'days' or info[2] == 'd':
                            fecha = info[0].shift(days=float(info[1]))
                            hoja.cell(row = i, column = j, value = fecha.format('DD/MM/YYYY'))
    archivo.save(file)


#''' Comprobado, estas funciones no sirven, es necesario abrir el archivo
# dentro de la funcion'''
# def find_test(word):
#     ''' Barre desde la celda A1 hasta la ultima escrita compara
#     los valores con la palabra y retorna el valor de la celda '''
#     for i in range (1, hoja.max_row, 1):
#         for j in range (1, hoja.max_column, 1):
#             info = hoja.cell(row = i, column = j, )
#             if info.value == word:
#                 goal = hoja.cell(row = i, column = j, )
#     return goal
#
#
# ''' Te dice en que celda se encuentra el valor buscado '''
# def o_pen(file, xlsheet, word):
#     ''' Abre Archivo '''
#     archivo = openpyxl.load_workbook(file)
#     hoja = archivo[xlsheet]
#     goalx = find_test(word)
#     return goalx
